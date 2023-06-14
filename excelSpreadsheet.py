"""
File: excelSpreadsheet.py
Author: Nikita Istomin
Creatin Date: 6/13/2023
Description: This file contains the ExcelSpreadsheet class which is used to open, create, and write to an Excel spreadsheet, 
    as well as perform other operations on data that require data from the Excel spreadsheet(like filtering out videos that are already in the spreadsheet).
"""

import time

import openpyxl
from openpyxl.styles import Font, PatternFill
from schemas.excelMatchEntry import ExcelMatchEntry, ExcelMatchEntryOrder

class ExcelSpreadsheet():

    # Colors for valorant ranks
    IRON_COLOR = "5f5f5f"
    BRONZE_COLOR = "bb9467"
    SILVER_COLOR = "c0c0c0"
    GOLD_COLOR = "ffd700"
    PLATINUM_COLOR = "00a7b4"
    DIAMOND_COLOR = "d381fc"
    ASCENDANT_COLOR = "30a46f"
    IMMORTAL_COLOR = "b7426c"
    RADIANT_COLOR = "f3de98"


    def __init__(self, filename : str):
        self.filename = filename
        self.wb = None


    def open(self) -> openpyxl.Workbook:
        """Open/create the excel file wiht the filename passed to the constructor

        Returns:
            openpyxl.Workbook: Workbook object
        """
        while True:
            print(f"Opening {self.filename}...")
            try:
                self.wb = openpyxl.load_workbook(self.filename)
            except FileNotFoundError:
                print(f"Error: File {self.filename} not found")
                print(f"Creating {self.filename}...")
                return self.create()
            # if file is open
            except PermissionError:
                secondsToWait = 5
                print(f"Error: File {self.filename} is open, please close.")
                print(f"Retrying in {secondsToWait} seconds...")
                time.sleep(1)
                for i in range(secondsToWait):
                    print(f"{secondsToWait - 1 - i}...")
                    time.sleep(1)
                continue # retry
            except Exception as e:
                raise e
            return self.wb
        


    def create(self) -> openpyxl.Workbook:
        """Create the excel file with the filename passed to the constructor

        Returns:
            openpyxl.Workbook: Workbook object
        """
        self.wb = openpyxl.Workbook()
        sheet = self.wb.active
        firstRow = ExcelMatchEntry.getPrintableTitlesInOrder()
        # Create first row
        sheet.append(firstRow)
        # Style the first row
        for cell in sheet[1]:
            # Make the first row bold
            cell.font = Font(bold=True, color='FFFFFF')
            # Change the background color of the first row
            cell.fill = PatternFill(start_color='538dd5', end_color='538dd5', fill_type='solid')
        self.wb.save(self.filename)
        return self.wb


    def writeMatcheToExcel(self, match : ExcelMatchEntry) -> None:
        """Write a match to the excel file

        Args:
            matche (ExcelMatchEntry): Match to write to the excel file

        Returns:
            None
        """
        self.writeMatchesToExcel([match])


    def writeMatchesToExcel(self, matches : list[ExcelMatchEntry]) -> None:
        """Write the matches to the excel file

        Args:
            matches (list[ExcelMatchEntry]): List of matches to write to the excel file

        Returns:
            None
        """
        if self.wb is None:
            print("Error: Workbook not open")
            return
        sheet = self.wb.active

        # Write the matches to the sheet
        for match in matches:
            sheet.append(match.asTuple())
            # Color in the match result cell
            matchResultCell = sheet[sheet.max_row][ExcelMatchEntryOrder.matchResult]
            if matchResultCell.value == "Win":
                # Green background
                matchResultCell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            elif matchResultCell.value == "Loss":
                # Red background
                matchResultCell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            else:
                # Grey background
                matchResultCell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
        self.wb.save(self.filename)


    def filterOutProcessedMatches(self, mp4_files : list[str]) -> list[str]:
        """Filter out videos that are already in the excel file
        
        Args:
            mp4_files (list[str]): List of MP4 file paths to filter
            EXCEL_FILE (str): Path to the excel file

        Returns:
            list[str]: List of paths to unprocessed videos
        """
        if self.wb is None:
            print("Error: Workbook not open")
            return
        sheet = self.wb.active

        # Get the match IDs from the excel file
        localPaths = [cell.value for cell in sheet["ABCDEFGHIJKLMNOPQRSTUVWXYZ"[ExcelMatchEntryOrder.localPath.value]][1:]]

        # Filter out the videos that are already in the excel file
        unprocessed_videos = [video for video in mp4_files if video not in localPaths]

        # Return the list of unprocessed videos
        return unprocessed_videos



# Testing
# if __name__ == "__main__":
#     import os
#     # Test the functions
#     filename = os.path.dirname(os.path.abspath(__file__)) + "\\test.xlsx"

#     excel = ExcelSpreadsheet(filename)

#     wb = excel.open()
#     sheet = wb.active

#     cell = sheet.cell(row=1, column=1)
#     cell.value = "Iron"
#     cell.fill = PatternFill(start_color=ExcelSpreadsheet.IRON_COLOR, end_color=ExcelSpreadsheet.IRON_COLOR, fill_type='solid')
#     cell = sheet.cell(row=2, column=1)
#     cell.value = "Bronze"
#     cell.fill = PatternFill(start_color=ExcelSpreadsheet.BRONZE_COLOR, end_color=ExcelSpreadsheet.BRONZE_COLOR, fill_type='solid')
#     cell = sheet.cell(row=3, column=1)
#     cell.value = "Silver"
#     cell.fill = PatternFill(start_color=ExcelSpreadsheet.SILVER_COLOR, end_color=ExcelSpreadsheet.SILVER_COLOR, fill_type='solid')
#     cell = sheet.cell(row=4, column=1)
#     cell.value = "Gold"
#     cell.fill = PatternFill(start_color=ExcelSpreadsheet.GOLD_COLOR, end_color=ExcelSpreadsheet.GOLD_COLOR, fill_type='solid')
#     cell = sheet.cell(row=5, column=1)
#     cell.value = "Platinum"
#     cell.fill = PatternFill(start_color=ExcelSpreadsheet.PLATINUM_COLOR, end_color=ExcelSpreadsheet.PLATINUM_COLOR, fill_type='solid')
#     cell = sheet.cell(row=6, column=1)
#     cell.value = "Diamond"
#     cell.fill = PatternFill(start_color=ExcelSpreadsheet.DIAMOND_COLOR, end_color=ExcelSpreadsheet.DIAMOND_COLOR, fill_type='solid')
#     cell = sheet.cell(row=7, column=1)
#     cell.value = "Ascendant"
#     cell.fill = PatternFill(start_color=ExcelSpreadsheet.ASCENDANT_COLOR, end_color=ExcelSpreadsheet.ASCENDANT_COLOR, fill_type='solid')
#     cell = sheet.cell(row=8, column=1)
#     cell.value = "Immortal"
#     cell.fill = PatternFill(start_color=ExcelSpreadsheet.IMMORTAL_COLOR, end_color=ExcelSpreadsheet.IMMORTAL_COLOR, fill_type='solid')
#     cell = sheet.cell(row=9, column=1)
#     cell.value = "Radiant"
#     cell.fill = PatternFill(start_color=ExcelSpreadsheet.RADIANT_COLOR, end_color=ExcelSpreadsheet.RADIANT_COLOR, fill_type='solid')

#     wb.save(filename)

#     localPaths = [cell.value for cell in sheet["ABCDEFGHIJKLMNOPQRSTUVWXYZ"[ExcelMatchEntryOrder.localPath.value]][1:]]
#     print(localPaths)